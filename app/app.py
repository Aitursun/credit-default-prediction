"""Credit scoring Streamlit app — ВКР КГИПИ 2026."""
import io
import json
import logging
import sys
import warnings
from datetime import datetime
from pathlib import Path

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))

import joblib
import matplotlib
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import shap
import streamlit as st
from sklearn.metrics import fbeta_score, precision_score, recall_score, roc_auc_score

from src.config import ID_COL, MODELS_DIR, PROC_DATA_DIR, TARGET_COL
from src.evaluation import business_metric, find_optimal_threshold

warnings.filterwarnings("ignore")
matplotlib.use("Agg")

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
    handlers=[logging.StreamHandler(sys.stdout)],
)
logger = logging.getLogger("credit_scoring")

# ─── Page config ─────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Кредитный скоринг | КГИПИ",
    page_icon="🏦",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─── Constants ────────────────────────────────────────────────────────────────
# Порядок по ROC-AUC из 5-fold CV (ноутбук 04):
# CatBoost 0.786 > EBM 0.785 > LGBM 0.778 > LogReg 0.771 > RF 0.757
MODEL_OPTIONS: dict[str, str] = {
    "CatBoost (рекомендуется)": "catboost",
    "EBM (InterpretML)": "ebm",
    "LightGBM": "lgbm",
    "Logistic Regression": "logreg",
    "Random Forest": "rf",
}

DISPLAY_MODES = ["Полный (для всех)", "Только для специалиста", "Только для аналитика"]

REQUIRED_BATCH_COLS = ["AMT_INCOME_TOTAL", "AMT_ANNUITY"]

FEATURE_EXPLANATIONS: dict[str, tuple[str, str, str, str]] = {
    # key: (Русское название, описание при риске HIGH, описание при риске LOW, единица)
    "EXT_SOURCE_MEAN":     ("Средний скоринг кредитных бюро",  "Низкий — плохая история погашений",        "Высокий — хорошая история погашений",       "0–1"),
    "EXT_SOURCE_2":        ("Скоринг бюро 2",                  "Низкий — плохая история",                  "Высокий — хорошая история",                 "0–1"),
    "EXT_SOURCE_3":        ("Скоринг бюро 3",                  "Низкий — плохая история",                  "Высокий — хорошая история",                 "0–1"),
    "EXT_SOURCE_1":        ("Скоринг бюро 1",                  "Низкий — плохая история",                  "Высокий — хорошая история",                 "0–1"),
    "EXT_SOURCE_MIN":      ("Минимальный скоринг бюро",        "Один из скорингов очень низкий",            "Все скоринги в норме",                      "0–1"),
    "EXT_SOURCE_MAX":      ("Максимальный скоринг бюро",       "Даже лучший скоринг низкий",               "Лучший скоринг высокий",                    "0–1"),
    "EXT_SOURCE_STD":      ("Разброс скорингов бюро",          "Большой разброс — нестабильность",         "Скоринги согласованны",                     ""),
    "ANNUITY_INCOME_RATIO":("Долговая нагрузка (платёж/доход)","Превышает норму 30% — высокий риск",       "В пределах нормы",                          "%"),
    "CREDIT_INCOME_RATIO": ("Отношение кредита к доходу",      "Кредит значительно превышает доход",       "Кредит соразмерен доходу",                  "×"),
    "ANNUITY_CREDIT_RATIO":("Доля платежа от кредита",         "Высокая доля — короткий срок погашения",   "Умеренная доля платежа",                    ""),
    "CREDIT_GOODS_RATIO":  ("Кредит к стоимости товара",       "Кредит превышает стоимость товара",        "Кредит соответствует стоимости",            ""),
    "AMT_INCOME_TOTAL":    ("Годовой доход",                   "Низкий доход повышает риск",               "Высокий доход снижает риск",                "сом"),
    "AMT_ANNUITY":         ("Ежемесячный платёж",              "Высокий платёж",                           "Умеренный платёж",                          "сом"),
    "AMT_GOODS_PRICE":     ("Сумма кредита",                   "Очень крупный кредит",                     "Умеренная сумма",                           "сом"),
    "DAYS_BIRTH":          ("Возраст заёмщика",                "Молодой возраст — статистически выше риск","Зрелый возраст — ниже риск",                "лет"),
    "YEARS_EMPLOYED":      ("Трудовой стаж",                   "Короткий стаж — нестабильная занятость",   "Длительный стаж — стабильная занятость",    "лет"),
    "IS_UNEMPLOYED":       ("Статус занятости",                "Безработный — значительно повышает риск",  "Трудоустроен",                              ""),
    "FLAG_OWN_REALTY_Y":   ("Наличие недвижимости",            "Нет недвижимости",                        "Есть недвижимость — снижает риск",          ""),
    "FLAG_OWN_CAR_Y":      ("Наличие автомобиля",              "Нет автомобиля",                          "Есть автомобиль — снижает риск",            ""),
    "CNT_CHILDREN":        ("Количество детей",                "Много детей — высокая финансовая нагрузка","Детей нет или мало",                        ""),
    "CNT_FAM_MEMBERS":     ("Размер семьи",                    "Большая семья — высокая нагрузка",         "Небольшая семья",                           ""),
    "INCOME_PER_PERSON":   ("Доход на члена семьи",            "Низкий доход на человека",                 "Достаточный доход на члена семьи",          "сом"),
    "CODE_GENDER_M":       ("Пол заёмщика",                    "Мужской",                                 "Женский",                                   ""),
    "AMT_REQ_CREDIT_BUREAU_TOTAL": ("Запросы в кредитное бюро","Много запросов — активно ищет кредиты",   "Запросов мало",                             "шт."),
    "DOCUMENT_COUNT":      ("Количество документов",           "Мало документов",                          "Полный пакет документов",                   "шт."),
    "BUREAU_LOAN_COUNT":   ("Кредитов в истории",              "Много кредитов",                           "Умеренное количество",                      "шт."),
    "BUREAU_ACTIVE_COUNT": ("Активных кредитов сейчас",        "Много активных кредитов",                  "Мало активных кредитов",                    "шт."),
    "BUREAU_ACTIVE_OVERDUE_SUM":("Сумма просрочек по активным","Есть просрочки — серьёзный риск",          "Просрочек нет",                             "сом"),
    "BB_STATUS_DPD_COUNT_SUM":  ("Случаев просрочки в истории","Много просрочек в прошлом",                "Просрочек нет",                             ""),
    "INS_DPD_MEAN":        ("Среднее дней просрочки",          "Высокая просрочка по рассрочкам",          "Платит вовремя",                            "дней"),
    "INS_OVERDUE_RATIO":   ("Доля просроченных платежей",      "Высокая доля просрочек",                   "Платит стабильно",                          "%"),
    "PREV_REFUSED_RATIO":  ("Доля отказов в прошлом",          "Часто отказывали — плохая история",        "Преимущественно одобряли",                  "%"),
    "POS_SK_DPD_DEF_MEAN": ("Дефолтные просрочки (POS)",       "Есть дефолтные просрочки",                 "Дефолтных просрочек нет",                   ""),
    "CC_SK_DPD_DEF_MEAN":  ("Дефолтные просрочки (карта)",     "Есть дефолтные просрочки по карте",        "Дефолтных просрочек нет",                   ""),
    "CC_UTILIZATION_MEAN": ("Утилизация кредитной карты",      "Высокая утилизация карты",                 "Умеренная утилизация",                      "%"),
    "REGION_RATING_CLIENT_W_CITY": ("Рейтинг региона",         "Неблагоприятный регион проживания",        "Благоприятный регион",                      ""),
    "DAYS_REGISTRATION_AGE_RATIO": ("Регистрация / возраст",   "Поздняя регистрация относительно возраста","Ранняя регистрация",                        ""),
    "DAYS_ID_PUBLISH_AGE_RATIO":   ("Смена документов / возраст","Поздняя смена документов",              "Документы в порядке",                       ""),
    "PHONE_TO_BIRTH_RATIO":        ("Смена телефона / возраст","Частая смена телефона",                    "Стабильный контакт",                        ""),
    "OWN_CAR_AGE":         ("Возраст автомобиля",              "Старый автомобиль",                        "Новый автомобиль",                          "лет"),
    "DEF_30_CNT_SOCIAL_CIRCLE":    ("Дефолты в окружении",     "Много дефолтов у знакомых",                "В окружении нет дефолтов",                  ""),
}

# ─── Caching & Loading ────────────────────────────────────────────────────────

@st.cache_resource
def load_model(name: str):
    path = MODELS_DIR / f"{name}.joblib"
    if not path.exists():
        logger.error("Файл модели не найден: %s", path)
        st.error(f"Файл модели не найден: {path}. Запустите ноутбук 04_train_models.ipynb.")
        st.stop()
    m = joblib.load(path)
    has_scaler = hasattr(m, "_scaler")
    logger.info("Модель загружена: %s  (%s)  scaler=%s", name.upper(), type(m).__name__, has_scaler)
    return m


def predict_proba_safe(model, X: pd.DataFrame) -> np.ndarray:
    """Apply _scaler if the model was trained with one, then return class-1 probabilities."""
    X_in = X.to_numpy(dtype=np.float32)
    if hasattr(model, "_scaler"):
        X_in = model._scaler.transform(X_in)
    return model.predict_proba(X_in)[:, 1]


@st.cache_data
def load_train_stats() -> tuple[pd.Series, list[str], pd.DataFrame]:
    """Returns (medians, feature_cols, X_background).
    Uses src.training.load_data() — same path as notebook 04."""
    from src.training import load_data
    X, _ = load_data()
    feature_cols = list(X.columns)
    medians = X.median()
    shap_cache = joblib.load(MODELS_DIR / "lgbm_shap.joblib")
    X_background: pd.DataFrame = shap_cache["X_sample"]
    return medians, feature_cols, X_background


_THRESHOLDS_PATH = PROC_DATA_DIR / "thresholds.json"


@st.cache_data
def compute_threshold(model_name: str, beta: float = 2.0) -> tuple[float, float]:
    """Load pre-computed threshold from thresholds.json (fast, instant).

    File is generated once by running:
        python -c "from app.app import _precompute_thresholds; _precompute_thresholds()"
    or via the admin button in the sidebar.
    Falls back to holdout computation if the file is missing.
    """
    if _THRESHOLDS_PATH.exists():
        data = json.loads(_THRESHOLDS_PATH.read_text())
        if model_name in data:
            entry = data[model_name]
            thr, f2 = float(entry["threshold"]), float(entry.get("f2", 0.0))
            logger.info("Порог [%s] загружен из файла: %.4f  F₂=%.4f", model_name.upper(), thr, f2)
            return thr, f2

    # Файл отсутствует или модели нет — вычисляем на holdout (быстро, секунды)
    logger.warning("thresholds.json не найден — вычисляю на holdout для %s", model_name.upper())
    from src.training import load_data
    X, y = load_data()
    y_arr = y.to_numpy() if hasattr(y, "to_numpy") else y
    _, X_val, _, y_val = __import__("sklearn.model_selection", fromlist=["train_test_split"]).train_test_split(
        X, y_arr, test_size=0.3, stratify=y_arr, random_state=42
    )
    m = load_model(model_name)
    probs = predict_proba_safe(m, X_val)
    threshold, fbeta = find_optimal_threshold(y_val, probs, beta=beta)
    logger.info("Порог [%s] holdout: %.4f  F₂=%.4f", model_name.upper(), threshold, fbeta)
    return float(threshold), float(fbeta)


def _scale_if_needed(model, X: pd.DataFrame | np.ndarray) -> np.ndarray:
    arr = X.to_numpy(dtype=np.float32) if isinstance(X, pd.DataFrame) else X.astype(np.float32)
    if hasattr(model, "_scaler"):
        arr = model._scaler.transform(arr)
    return arr


@st.cache_resource
def get_shap_explainer(_model, _X_background: pd.DataFrame, model_name: str):
    """Build SHAP explainer. model_name used only for cache-key differentiation."""
    try:
        bg = _scale_if_needed(_model, _X_background[:500])
        if model_name == "logreg":
            return shap.LinearExplainer(_model, bg)
        elif model_name == "ebm":
            return None  # use EBM's own explain_local
        else:
            return shap.TreeExplainer(
                _model, _scale_if_needed(_model, _X_background[:200]),
                feature_perturbation="interventional",
            )
    except Exception:
        try:
            return shap.TreeExplainer(_model)
        except Exception:
            return None


# ─── Feature preparation ─────────────────────────────────────────────────────

def prepare_single_input(form: dict, medians: pd.Series, feature_cols: list[str]) -> pd.DataFrame:
    """Map form values → 209-feature DataFrame row."""
    row = medians.copy()

    income = max(form["income"], 1)
    annuity = form["annuity"]
    credit = max(form["credit"], 1)
    age = form["age"]
    years_employed = form["years_employed"]
    is_unemployed = form["is_unemployed"]

    # Base application fields
    row["DAYS_BIRTH"] = -age * 365
    row["AMT_INCOME_TOTAL"] = income
    row["AMT_ANNUITY"] = annuity
    row["CNT_CHILDREN"] = form["children"]
    row["CNT_FAM_MEMBERS"] = form["family_size"]
    row["CODE_GENDER_M"] = 1 if form["gender"] == "Мужской" else 0
    row["FLAG_OWN_CAR_Y"] = int(form["own_car"])
    row["FLAG_OWN_REALTY_Y"] = int(form["own_realty"])

    # Employment
    if is_unemployed:
        row["IS_UNEMPLOYED"] = 1
        row["YEARS_EMPLOYED"] = 0.0
    else:
        row["IS_UNEMPLOYED"] = 0
        row["YEARS_EMPLOYED"] = years_employed

    # External sources
    for i in range(1, 4):
        val = form.get(f"ext_source_{i}")
        col = f"EXT_SOURCE_{i}"
        miss_col = f"EXT_SOURCE_{i}_MISSING"
        if val is not None:
            row[col] = val
            if miss_col in row.index:
                row[miss_col] = 0
        # else keep median and missing flag as-is

    # Сумма кредита → AMT_GOODS_PRICE (proxy) и AMT_CREDIT если присутствует в фичах
    row["AMT_GOODS_PRICE"] = credit
    if "AMT_CREDIT" in row.index:
        row["AMT_CREDIT"] = credit

    # Производные коэффициенты
    row["CREDIT_INCOME_RATIO"] = credit / income
    row["ANNUITY_INCOME_RATIO"] = annuity / income
    row["CREDIT_GOODS_RATIO"] = 1.0          # AMT_CREDIT / AMT_GOODS_PRICE = 1 (прокси)
    row["ANNUITY_CREDIT_RATIO"] = annuity / credit

    ext_vals = [row[f"EXT_SOURCE_{i}"] for i in range(1, 4) if not np.isnan(row.get(f"EXT_SOURCE_{i}", np.nan))]
    if ext_vals:
        row["EXT_SOURCE_MEAN"] = float(np.mean(ext_vals))
        row["EXT_SOURCE_MIN"] = float(np.min(ext_vals))
        row["EXT_SOURCE_MAX"] = float(np.max(ext_vals))
        row["EXT_SOURCE_STD"] = float(np.std(ext_vals)) if len(ext_vals) > 1 else 0.0

    fam = max(form["family_size"], 1)
    row["INCOME_PER_PERSON"] = income / fam

    # Семейный статус → one-hot (NAME_FAMILY_STATUS_*)
    married = form.get("married", True)
    for col in ["NAME_FAMILY_STATUS_Married", "NAME_FAMILY_STATUS_Single / not married",
                "NAME_FAMILY_STATUS_Civil marriage", "NAME_FAMILY_STATUS_Separated",
                "NAME_FAMILY_STATUS_Widow"]:
        if col in row.index:
            row[col] = 0
    if "NAME_FAMILY_STATUS_Married" in row.index:
        row["NAME_FAMILY_STATUS_Married"] = 1 if married else 0
    if "NAME_FAMILY_STATUS_Single / not married" in row.index:
        row["NAME_FAMILY_STATUS_Single / not married"] = 0 if married else 1

    return pd.DataFrame([row[feature_cols]], columns=feature_cols)


def prepare_batch_input(
    uploaded_df: pd.DataFrame,
    medians: pd.Series,
    feature_cols: list[str],
) -> pd.DataFrame:
    """Map batch upload columns → 209-feature DataFrame."""
    out = pd.DataFrame(
        np.tile(medians[feature_cols].values, (len(uploaded_df), 1)),
        columns=feature_cols,
    )

    # Copy any matching columns directly
    for col in feature_cols:
        if col in uploaded_df.columns:
            out[col] = uploaded_df[col].values

    # Handle human-friendly alias: AGE_YEARS → DAYS_BIRTH
    if "AGE_YEARS" in uploaded_df.columns and "DAYS_BIRTH" in feature_cols:
        out["DAYS_BIRTH"] = -uploaded_df["AGE_YEARS"].values * 365

    # Handle AMT_CREDIT → proxy for AMT_GOODS_PRICE + ratios
    if "AMT_CREDIT" in uploaded_df.columns:
        credit = uploaded_df["AMT_CREDIT"].clip(lower=1).values
        income = uploaded_df.get("AMT_INCOME_TOTAL", pd.Series([medians["AMT_INCOME_TOTAL"]] * len(uploaded_df))).clip(lower=1).values
        annuity = uploaded_df.get("AMT_ANNUITY", pd.Series([medians["AMT_ANNUITY"]] * len(uploaded_df))).values

        out["AMT_GOODS_PRICE"] = credit
        out["CREDIT_INCOME_RATIO"] = credit / income
        out["ANNUITY_INCOME_RATIO"] = annuity / income
        out["CREDIT_GOODS_RATIO"] = 1.0
        out["ANNUITY_CREDIT_RATIO"] = annuity / credit

    # Recompute EXT_SOURCE derived cols if raw cols are present
    ext_present = [f"EXT_SOURCE_{i}" in uploaded_df.columns for i in range(1, 4)]
    if any(ext_present):
        ext_df = pd.DataFrame({
            f"EXT_SOURCE_{i}": uploaded_df[f"EXT_SOURCE_{i}"].values
            if f"EXT_SOURCE_{i}" in uploaded_df.columns
            else np.full(len(uploaded_df), medians.get(f"EXT_SOURCE_{i}", np.nan))
            for i in range(1, 4)
        })
        out["EXT_SOURCE_MEAN"] = ext_df.mean(axis=1).values
        out["EXT_SOURCE_MIN"] = ext_df.min(axis=1).values
        out["EXT_SOURCE_MAX"] = ext_df.max(axis=1).values
        out["EXT_SOURCE_STD"] = ext_df.std(axis=1).values

    # Code gender
    if "CODE_GENDER" in uploaded_df.columns:
        out["CODE_GENDER_M"] = (uploaded_df["CODE_GENDER"].str.upper() == "M").astype(float).values

    # Boolean Y/N columns
    for src, dst in [("FLAG_OWN_CAR", "FLAG_OWN_CAR_Y"), ("FLAG_OWN_REALTY", "FLAG_OWN_REALTY_Y")]:
        if src in uploaded_df.columns and dst in feature_cols:
            out[dst] = (uploaded_df[src].astype(str).str.upper() == "Y").astype(float).values

    return out.astype(float)


# ─── SHAP computation ─────────────────────────────────────────────────────────

def compute_shap_values(
    model,
    model_name: str,
    X_input: pd.DataFrame,
    X_background: pd.DataFrame,
) -> tuple[np.ndarray | None, float | None]:
    """Returns (shap_values_1d, expected_value) or (None, None) on failure."""
    explainer = get_shap_explainer(model, X_background, model_name)
    if explainer is None:
        return None, None
    try:
        X_in = _scale_if_needed(model, X_input)
        shap_exp = explainer(X_in)
        vals = shap_exp.values
        if vals.ndim == 3:
            vals = vals[:, :, 1]
        ev = float(shap_exp.base_values[0]) if vals.ndim == 2 else float(shap_exp.base_values)
        if isinstance(ev, np.ndarray):
            ev = float(ev.flat[0])
        return vals[0], ev
    except Exception:
        return None, None


def get_ebm_local_importance(model, X_input: pd.DataFrame) -> tuple[np.ndarray | None, list[str] | None]:
    """Extract per-feature scores from EBM explain_local."""
    try:
        local_exp = model.explain_local(X_input, name="local")
        data = local_exp.data(0)
        scores = np.array(data["scores"])
        names = list(data["names"])
        return scores, names
    except Exception:
        return None, None


# ─── UI helpers ───────────────────────────────────────────────────────────────

# Жёсткие бизнес-правила — срабатывают независимо от ML-модели.
# Важно: в датасете AMT_INCOME_TOTAL — ежемесячный доход (тот же масштаб что AMT_ANNUITY),
# поэтому DTI = annuity / income сравнивается напрямую (норма ≤ 30%).
def check_business_rules(income: float, credit: float, annuity: float) -> list[str]:
    """Return list of violated hard-rule descriptions. Empty → no violations."""
    violations: list[str] = []
    if income <= 0:
        return violations
    dti = annuity / income          # = ANNUITY_INCOME_RATIO (оба в месяц)
    cti = credit / income           # кредит к месячному доходу
    if dti >= 1.0:
        violations.append(
            f"Платёж {annuity:,.0f} ≥ дохода {income:,.0f} сом "
            f"(долговая нагрузка {dti:.0%} — невозможно обслуживать)"
        )
    elif dti >= 0.80:
        violations.append(
            f"Платёж {annuity:,.0f} составляет {dti:.0%} дохода {income:,.0f} сом "
            f"(критически высокая нагрузка, норма ≤ 30%)"
        )
    if cti > 240:
        violations.append(
            f"Кредит {credit:,.0f} = {cti:.0f}× месячного дохода "
            f"(норма: до 240×, т.е. 20 лет)"
        )
    return violations


# ─── Зоны риска (Базель II IRB, EBA GL/2020/06, международная практика) ─────
#
# Зоны определяются ОТНОСИТЕЛЬНО порога каждой модели.
# Решает проблему разных шкал вероятностей:
#   CatBoost порог ~0.52, EBM порог ~0.09, LightGBM порог ~0.58
#
# Ниже порога:  closeness = prob / threshold  → [0, 1)
#   < 0.25  → Очень низкий
#   < 0.65  → Низкий
#   < 1.00  → Средний (приближается к порогу)
#
# Выше порога:  excess = (prob - thr) / (1 - thr)  → [0, 1)
#   < 0.10  → Пограничный  (чуть выше порога)
#   < 0.50  → Высокий
#   ≥ 0.50  → Очень высокий

ZONE_DATA = [
    # (код,         label_ru,          цвет,     label_решения,         иконка)
    ("very_low",   "Очень низкий",  "#1a7a3c", "АВТО-ОДОБРЕНИЕ",      "✅"),
    ("low",        "Низкий",        "#28a745", "ОДОБРИТЬ",             "✅"),
    ("medium",     "Средний",       "#6db33f", "ОДОБРИТЬ С УСЛОВИЕМ",  "✅"),
    ("borderline", "Пограничный",   "#e6a817", "РУЧНАЯ ПРОВЕРКА",      "⚠️"),
    ("high",       "Высокий",       "#e07320", "УСЛОВНЫЙ ОТКАЗ",       "⛔"),
    ("very_high",  "Очень высокий", "#dc3545", "АВТО-ОТКАЗ",           "❌"),
]


def get_zone(prob: float, threshold: float = 0.5) -> tuple[str, str, str, str, str]:
    """Определяет зону риска относительно порога конкретной модели."""
    thr = max(threshold, 1e-6)
    if prob >= thr:
        excess = (prob - thr) / max(1.0 - thr, 1e-6)
        idx = 3 if excess < 0.20 else (4 if excess < 0.50 else 5)
    else:
        closeness = prob / thr
        idx = 0 if closeness < 0.25 else (1 if closeness < 0.65 else 2)
    return tuple(ZONE_DATA[idx])

def risk_level_label(prob: float, threshold: float) -> str:
    return get_zone(prob, threshold)[1]


def show_gauge(prob: float, threshold: float) -> None:
    _, _, zone_color, _, _ = get_zone(prob, threshold)

    # Границы зон строятся из тех же коэффициентов что get_zone(),
    # динамически от реального порога каждой модели.
    t = threshold
    b1 = t * 0.25 * 100                      # very_low / low
    b2 = t * 0.65 * 100                      # low / medium
    b3 = t * 100                              # medium / borderline = порог
    b4 = (t + 0.20 * (1 - t)) * 100          # borderline / high
    b5 = (t + 0.50 * (1 - t)) * 100          # high / very_high

    fig = go.Figure(go.Indicator(
        mode="gauge+number+delta",
        value=prob * 100,
        number={"suffix": "%", "font": {"size": 40}},
        title={
            "text": "Риск невозврата<br><span style='font-size:12px;color:gray'>Вероятность дефолта</span>"
        },
        delta={
            "reference": threshold * 100,
            "valueformat": ".1f",
            "increasing": {"color": "red"},
            "decreasing": {"color": "green"},
        },
        gauge={
            "axis": {
                "range": [0, 100], "ticksuffix": "%",
                "tickvals": [0, round(b1), round(b2), round(b3), round(b4), round(b5), 100],
            },
            "bar": {"color": zone_color, "thickness": 0.3},
            "steps": [
                {"range": [0,    b1],  "color": "#c3e6cb"},  # очень низкий
                {"range": [b1,   b2],  "color": "#d4edda"},  # низкий
                {"range": [b2,   b3],  "color": "#e8f5d0"},  # средний
                {"range": [b3,   b4],  "color": "#fff3cd"},  # пограничный
                {"range": [b4,   b5],  "color": "#fde8c8"},  # высокий
                {"range": [b5, 100],   "color": "#f8d7da"},  # очень высокий
            ],
            "threshold": {
                "line": {"color": "black", "width": 3},
                "thickness": 0.75,
                "value": threshold * 100,
            },
        },
    ))
    fig.update_layout(height=280, margin=dict(t=60, b=0, l=20, r=20))
    st.plotly_chart(fig, use_container_width=True)
    st.caption(
        f"Зоны (от порога {threshold:.1%}): "
        f"0–{b1:.0f}% Очень низкий · {b1:.0f}–{b2:.0f}% Низкий · "
        f"{b2:.0f}–{b3:.0f}% Средний · {b3:.0f}–{b4:.0f}% Пограничный · "
        f"{b4:.0f}–{b5:.0f}% Высокий · {b5:.0f}–100% Очень высокий."
    )


def show_factors(
    shap_vals: np.ndarray,
    feature_cols: list[str],
    X_row: pd.DataFrame,
    mode: str,
) -> None:
    """Split top SHAP contributors into risk / protective columns."""
    pairs = sorted(
        zip(feature_cols, shap_vals, X_row.iloc[0].values),
        key=lambda x: abs(x[1]),
        reverse=True,
    )
    risk_items = [(f, s, v) for f, s, v in pairs if s > 0][:5]
    safe_items = [(f, s, v) for f, s, v in pairs if s < 0][:5]

    col_risk, col_safe = st.columns(2)

    def _render_item(col, feat: str, shap_val: float, feat_val: float, icon: str) -> None:
        info = FEATURE_EXPLANATIONS.get(feat)
        if info:
            rus_name, high_desc, low_desc, unit = info
            desc = high_desc if shap_val > 0 else low_desc
        else:
            rus_name = feat
            desc = ""
            unit = ""

        val_str = f"{feat_val:.3g} {unit}".strip() if not np.isnan(feat_val) else "—"
        with col:
            st.markdown(f"**{icon} {rus_name}**")
            if desc:
                st.markdown(f"<span style='color:#555;font-size:13px'>{desc}</span>", unsafe_allow_html=True)
            if mode in ("Полный (для всех)", "Только для аналитика"):
                st.caption(f"{feat} = {val_str} · SHAP: {shap_val:+.3f}")
            st.markdown("---")

    with col_risk:
        st.markdown("### 🔴 Факторы риска")
        for feat, shap_val, feat_val in risk_items:
            _render_item(col_risk, feat, shap_val, feat_val, "🔴")
        if not risk_items:
            st.info("Значимых факторов риска не выявлено.")

    with col_safe:
        st.markdown("### 🟢 Защитные факторы")
        for feat, shap_val, feat_val in safe_items:
            _render_item(col_safe, feat, shap_val, feat_val, "🟢")
        if not safe_items:
            st.info("Защитных факторов не выявлено.")


def show_waterfall(shap_vals: np.ndarray, expected_value: float, feature_cols: list[str], X_row: pd.DataFrame) -> None:
    shap_exp = shap.Explanation(
        values=shap_vals,
        base_values=expected_value,
        data=X_row.iloc[0].values,
        feature_names=feature_cols,
    )
    col1, col2 = st.columns([3, 2])
    with col1:
        st.caption("SHAP Waterfall Plot — вклад каждой переменной в итоговую вероятность")
        plt.figure()
        shap.plots.waterfall(shap_exp, max_display=14, show=False)
        st.pyplot(plt.gcf(), use_container_width=True)
        plt.close("all")

    with col2:
        st.caption("Расшифровка переменных:")
        top_pairs = sorted(zip(feature_cols, shap_vals), key=lambda x: abs(x[1]), reverse=True)[:12]
        table = []
        for feat, val in top_pairs:
            rus = FEATURE_EXPLANATIONS.get(feat, (feat,))[0]
            table.append({
                "Переменная": feat,
                "Название": rus,
                "Влияние": "⬆️ Риск" if val > 0 else "⬇️ Риск",
                "SHAP": f"{val:+.3f}",
            })
        st.dataframe(pd.DataFrame(table), hide_index=True, use_container_width=True)


def show_quick_stats(income: float, credit: float, annuity: float, term: int = 0) -> None:
    if income <= 0:
        return
    # AMT_INCOME_TOTAL и AMT_ANNUITY — одного масштаба (ежемесячные значения),
    # поэтому DTI = annuity / income, остаток = income − annuity (без деления на 12).
    dti = annuity / income
    cti = credit / income        # кредит / месячный доход (нормально до 240× = 20 лет)
    remainder = income - annuity  # остаток месячного дохода после платежа

    cols = st.columns(4) if term > 0 else st.columns(3)
    c1, c2, c3 = cols[0], cols[1], cols[2]

    c1.metric(
        "💳 Долговая нагрузка",
        f"{dti:.1%}",
        delta="норма ✅" if dti < 0.30 else "превышение ⚠️",
        delta_color="normal" if dti < 0.30 else "inverse",
        help="Отношение ежемесячного платежа к месячному доходу. Норма: до 30%.",
    )
    c1.caption(f"ANNUITY_INCOME_RATIO = {dti:.4f}")

    c2.metric(
        "📊 Кредит к месячному доходу",
        f"{cti:.0f}×",
        delta="норма ✅" if cti <= 240 else "высокое ⚠️",
        delta_color="normal" if cti <= 240 else "inverse",
        help="Во сколько месячных доходов обходится кредит. Норма: до 240× (20 лет).",
    )
    c2.caption(f"CREDIT_INCOME_RATIO = {cti:.4f}")

    c3.metric(
        "💰 Остаток после платежа",
        f"{remainder:,.0f} сом/мес",
        delta="положительный ✅" if remainder > 0 else "отрицательный ❌",
        delta_color="normal" if remainder > 0 else "inverse",
        help="Сколько остаётся на жизнь после выплаты кредита (из месячного дохода).",
    )
    c3.caption("AMT_INCOME_TOTAL − AMT_ANNUITY")

    if term > 0:
        total_paid = annuity * term
        overpay = total_paid - credit
        cols[3].metric(
            "📅 Всего выплат",
            f"{total_paid:,.0f} сом",
            delta=f"переплата {overpay:,.0f}" if overpay > 0 else "без переплаты",
            delta_color="inverse" if overpay > 0 else "normal",
            help=f"Сумма всех платежей за {term} мес. без учёта процентной ставки.",
        )
        cols[3].caption(f"ANNUITY_CREDIT_RATIO = {annuity/max(credit,1):.4f}")


def show_recommendation(prob: float, threshold: float) -> None:
    st.markdown("### 💬 Рекомендации кредитному специалисту")
    zone_code = get_zone(prob, threshold)[0]

    if zone_code == "very_low":
        st.success(
            "**Очень низкий риск — авто-одобрение.**  \n"
            "Вероятность дефолта < 15%. Согласно международной практике (Базель II, PD < 5%) "
            "заёмщик относится к наивысшей кредитной категории.  \n\n"
            "Дополнительно можно предложить:\n"
            "- Увеличение лимита или срока кредитования\n"
            "- Сниженную процентную ставку"
        )
    elif zone_code == "low":
        st.success(
            "**Низкий риск — одобрение на стандартных условиях.**  \n"
            "Вероятность дефолта 15–35%. Соответствует категории «Удовлетворительный» "
            "по EBA Guidelines on Loan Origination (EBA/GL/2020/06).  \n\n"
            "Стандартный пакет документов, стандартная ставка."
        )
    elif zone_code == "medium":
        st.success(
            "**Средний риск — одобрение с дополнительной проверкой.**  \n"
            "Вероятность дефолта 35–50%. Согласно Basel II IRB, PD 15–30% "
            "требует усиленного мониторинга.  \n\n"
            "Рекомендуется:\n"
            "- Проверить источник и стабильность дохода\n"
            "- Рассмотреть привлечение поручителя\n"
            "- Установить стандартную или повышенную ставку"
        )
    elif zone_code == "borderline":
        st.warning(
            "**Пограничная зона — направить на ручную проверку кредитного офицера.**  \n"
            "Вероятность дефолта 50–65% — вблизи порога отсечения модели. "
            "Согласно Базель II (judgmental override), случаи вблизи границы "
            "не могут приниматься автоматически и требуют оценки специалиста.  \n\n"
            "Офицер должен проверить:\n"
            "- Подтверждённый доход (справка, выписка с р/с)\n"
            "- Историю платежей по действующим обязательствам\n"
            "- Возможность снизить сумму кредита на 20–30% или запросить залог"
        )
    elif zone_code == "high":
        st.error(
            "**Высокий риск — условный отказ.**  \n"
            "Вероятность дефолта 65–80%. Соответствует категории «Субстандартный» "
            "по классификации банковских активов МВФ/Базель II.  \n\n"
            "Отказ может быть пересмотрен при:\n"
            "- Предоставлении ликвидного залога (≥ 150% суммы кредита)\n"
            "- Привлечении платёжеспособного поручителя\n"
            "- Уменьшении суммы кредита более чем на 40%"
        )
    else:  # very_high
        st.error(
            "**Очень высокий риск — авто-отказ.**  \n"
            "Вероятность дефолта > 80%. Согласно Базель II и требованиям "
            "Национального банка КР, кредитование заёмщиков с PD > 70% "
            "без обеспечения не рекомендовано.  \n\n"
            "Рекомендуется отказать. Альтернатив в данном случае нет."
        )


# ─── Excel export ─────────────────────────────────────────────────────────────

def generate_excel_report(
    result_df: pd.DataFrame,
    threshold: float,
    model_name: str,
) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()

    green_fill = PatternFill(start_color="C6EFCE", fill_type="solid")
    red_fill   = PatternFill(start_color="FFC7CE", fill_type="solid")
    hdr_fill   = PatternFill(start_color="1F4E79", fill_type="solid")
    hdr_font   = Font(color="FFFFFF", bold=True)
    gray_font  = Font(color="888888", size=9)

    # Sheet 1: Results
    ws1 = wb.active
    ws1.title = "Результаты"

    display_cols = {
        "Возраст (лет)": "age_display",
        "Годовой доход": "AMT_INCOME_TOTAL",
        "Кредит": "AMT_GOODS_PRICE",
        "Платёж/мес": "AMT_ANNUITY",
        "Риск (%)": "default_prob_pct",
        "Уровень риска": "risk_level",
        "Решение": "decision",
    }

    tech_row = ["SK_ID_CURR"] + ["AGE_YEARS", "AMT_INCOME_TOTAL", "AMT_GOODS_PRICE",
                                  "AMT_ANNUITY", "default_prob", "risk_level", "decision"]
    rus_row  = ["ID заявки"] + list(display_cols.keys())

    # Header row 1 (Russian)
    ws1.append(rus_row)
    for cell in ws1[1]:
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")

    # Header row 2 (technical, gray)
    ws1.append(tech_row)
    for cell in ws1[2]:
        cell.font = gray_font

    export_cols = ["SK_ID_CURR", "AMT_INCOME_TOTAL", "AMT_GOODS_PRICE",
                   "AMT_ANNUITY", "default_prob_pct", "risk_level", "decision"]
    avail = [c for c in export_cols if c in result_df.columns]
    for row_vals in result_df[avail].itertuples(index=False):
        ws1.append(list(row_vals))

    # Colour rows
    for row in ws1.iter_rows(min_row=3):
        decision_cell = row[-1]
        fill = red_fill if decision_cell.value == "Отказать" else green_fill
        for cell in row:
            cell.fill = fill

    # Sheet 2: Stats
    ws2 = wb.create_sheet("Статистика")
    n_approved = (result_df["decision"] == "Одобрить").sum()
    n_rejected = (result_df["decision"] == "Отказать").sum()
    stats = [
        ["Показатель", "Значение"],
        ["Всего заявок", len(result_df)],
        ["Одобрено", f"{n_approved} ({n_approved/len(result_df):.1%})"],
        ["Отказано",  f"{n_rejected} ({n_rejected/len(result_df):.1%})"],
        ["Средний риск", f"{result_df['default_prob'].mean():.1%}"],
        ["Медиана риска", f"{result_df['default_prob'].median():.1%}"],
        ["Дата оценки", datetime.now().strftime("%d.%m.%Y %H:%M")],
    ]
    for s in stats:
        ws2.append(s)

    # Sheet 3: Model info
    ws3 = wb.create_sheet("О модели")
    for row in [
        ["Параметр", "Значение"],
        ["Модель", model_name.upper()],
        ["Порог отсечения", f"{threshold:.4f}"],
        ["Метрика оптимизации", "F₂-score (β=2)"],
        ["Принцип", "Recall важнее Precision — снижение пропущенных дефолтов"],
        ["Дата оценки", datetime.now().strftime("%d.%m.%Y %H:%M")],
    ]:
        ws3.append(row)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ─── Template CSV ─────────────────────────────────────────────────────────────

def make_template_csv() -> str:
    template = pd.DataFrame({
        "AMT_INCOME_TOTAL": [150000, 90000],
        "AMT_CREDIT":       [450000, 300000],
        "AMT_ANNUITY":      [22500,  18000],
        "AGE_YEARS":        [35, 28],
        "CODE_GENDER":      ["F", "M"],
        "FLAG_OWN_CAR":     ["N", "Y"],
        "FLAG_OWN_REALTY":  ["Y", "N"],
        "CNT_CHILDREN":     [0, 1],
        "CNT_FAM_MEMBERS":  [2, 3],
        "YEARS_EMPLOYED":   [5, 1],
        "IS_UNEMPLOYED":    [0, 0],
        "EXT_SOURCE_1":     [0.60, 0.30],
        "EXT_SOURCE_2":     [0.75, 0.35],
        "EXT_SOURCE_3":     [0.65, 0.40],
    })
    return template.to_csv(index=False, encoding="utf-8-sig")


# ─── Sidebar ──────────────────────────────────────────────────────────────────

with st.sidebar:
    st.title("🏦 Кредитный скоринг")
    st.caption("ВКР · КГИПИ · 2026")
    st.markdown("---")

    st.markdown("**── Модель ──**")
    model_label = st.radio(
        "Выберите модель",
        list(MODEL_OPTIONS.keys()),
        label_visibility="collapsed",
    )
    selected_model_name = MODEL_OPTIONS[model_label]

    st.markdown("---")
    st.markdown("**── Режим отображения ──**")
    display_mode = st.radio(
        "Режим",
        DISPLAY_MODES,
        label_visibility="collapsed",
    )

    st.markdown("---")
    st.markdown("**── О системе ──**")
    st.info(
        "ℹ️ Система автоматически оценивает вероятность невозврата кредита. "
        "Решение носит **рекомендательный** характер."
    )
    st.download_button(
        "📥 Скачать шаблон CSV",
        data=make_template_csv(),
        file_name="template_scoring.csv",
        mime="text/csv",
    )



# ─── Load resources ───────────────────────────────────────────────────────────

with st.spinner("Загрузка модели и данных…"):
    medians, feature_cols, X_background = load_train_stats()
    model = load_model(selected_model_name)
    threshold, _ = compute_threshold(selected_model_name)

with st.sidebar:
    st.markdown("---")
    st.markdown("**── Порог отсечения ──**")
    st.metric(
        f"{selected_model_name.upper()} · F₂ (β=2)",
        f"{threshold:.4f}",
        help="Вероятность, при которой модель рекомендует отказ. Оптимизирован по F₂-мере.",
    )

show_mode_specialist  = display_mode in ("Полный (для всех)", "Только для специалиста")
show_mode_analyst     = display_mode in ("Полный (для всех)", "Только для аналитика")

# ─── Tabs ────────────────────────────────────────────────────────────────────

tab1, tab2 = st.tabs(["🔍 Проверить заёмщика", "📋 Загрузить список заявок"])

# ═══════════════════════════════════════════════════════════════════════════════
#  TAB 1 — Single borrower
# ═══════════════════════════════════════════════════════════════════════════════
with tab1:
    st.header("Оценка отдельного заёмщика")

    # ── Дефолтные значения session_state (первый запуск) ─────────────────────
    _DEFAULTS: dict = dict(
        age_input=35, gender_input="Женский", married_input=True,
        children_input=0, family_size_input=2, car_input=False, realty_input=False,
        income_input=150_000, credit_input=450_000, term_input=24, annuity_input=18_750,
        emp_input=3, unemp_input=False,
        use_ext1_input=True, ext1_input=float(round(medians["EXT_SOURCE_1"], 2)),
        use_ext2_input=True, ext2_input=float(round(medians["EXT_SOURCE_2"], 2)),
        use_ext3_input=True, ext3_input=float(round(medians["EXT_SOURCE_3"], 2)),
    )
    for _k, _v in _DEFAULTS.items():
        if _k not in st.session_state:
            st.session_state[_k] = _v

    # ── Пресеты (6 зон риска) ─────────────────────────────────────────────────
    PRESETS: dict[str, dict] = {
        "✅ Очень\nнизкий": dict(
            age_input=50, gender_input="Женский", married_input=True,
            children_input=0, family_size_input=2, car_input=True, realty_input=True,
            income_input=320_000, credit_input=250_000, term_input=25, annuity_input=10_000,
            emp_input=18, unemp_input=False,
            use_ext1_input=True, ext1_input=0.88,
            use_ext2_input=True, ext2_input=0.90,
            use_ext3_input=True, ext3_input=0.85,
        ),
        "✅ Низкий": dict(
            age_input=40, gender_input="Женский", married_input=True,
            children_input=0, family_size_input=2, car_input=False, realty_input=True,
            income_input=180_000, credit_input=250_000, term_input=20, annuity_input=12_500,
            emp_input=7, unemp_input=False,
            use_ext1_input=True, ext1_input=0.55,
            use_ext2_input=True, ext2_input=0.58,
            use_ext3_input=True, ext3_input=0.53,
        ),
        "✅ Средний": dict(
            age_input=34, gender_input="Женский", married_input=True,
            children_input=1, family_size_input=3, car_input=False, realty_input=False,
            income_input=120_000, credit_input=280_000, term_input=16, annuity_input=18_000,
            emp_input=3, unemp_input=False,
            use_ext1_input=True, ext1_input=0.42,
            use_ext2_input=True, ext2_input=0.44,
            use_ext3_input=True, ext3_input=0.40,
        ),
        "⚠️ Погранич-\nный": dict(
            age_input=28, gender_input="Женский", married_input=False,
            children_input=0, family_size_input=1, car_input=False, realty_input=False,
            income_input=90_000, credit_input=270_000, term_input=15, annuity_input=18_000,
            emp_input=2, unemp_input=False,
            use_ext1_input=True, ext1_input=0.32,
            use_ext2_input=True, ext2_input=0.35,
            use_ext3_input=True, ext3_input=0.30,
        ),
        "⛔ Высокий": dict(
            age_input=25, gender_input="Мужской", married_input=False,
            children_input=1, family_size_input=2, car_input=False, realty_input=False,
            income_input=58_000, credit_input=280_000, term_input=12, annuity_input=23_000,
            emp_input=1, unemp_input=False,
            use_ext1_input=True, ext1_input=0.12,
            use_ext2_input=True, ext2_input=0.14,
            use_ext3_input=True, ext3_input=0.11,
        ),
        "❌ Очень\nвысокий": dict(
            age_input=21, gender_input="Мужской", married_input=False,
            children_input=2, family_size_input=3, car_input=False, realty_input=False,
            income_input=35_000, credit_input=240_000, term_input=15, annuity_input=16_000,
            emp_input=0, unemp_input=True,
            use_ext1_input=True, ext1_input=0.05,
            use_ext2_input=True, ext2_input=0.06,
            use_ext3_input=True, ext3_input=0.05,
        ),
    }

    st.caption("Примеры для всех зон риска:")
    preset_cols = st.columns(6)
    for col, (label, values) in zip(preset_cols, PRESETS.items()):
        if col.button(label, use_container_width=True):
            st.session_state.update(values)
            st.session_state.pop("tab1_result", None)
            st.rerun()

    st.markdown("---")

    # ── Callbacks ─────────────────────────────────────────────────────────────
    def _sync_family() -> None:
        n       = int(st.session_state.get("children_input", 0))
        married = bool(st.session_state.get("married_input", True))
        min_fam = n + (2 if married else 1)
        if int(st.session_state.get("family_size_input", min_fam)) < min_fam:
            st.session_state.family_size_input = min_fam

    def _sync_annuity() -> None:
        c = float(st.session_state.get("credit_input", 450_000))
        t = int(st.session_state.get("term_input", 24))
        st.session_state.annuity_input = max(1, int(c / max(t, 1)))

    # ── Личные данные ─────────────────────────────────────────────────────────
    with st.expander("▼ Личные данные", expanded=True):
        # Строка 1: числовые поля и радио
        r1c1, r1c2, r1c3, r1c4 = st.columns(4)
        age      = r1c1.number_input("Возраст", 18, 75, step=1, key="age_input")
        gender   = r1c2.radio("Пол", ["Мужской", "Женский"], horizontal=True, key="gender_input")
        children = r1c3.number_input("Кол-во детей", 0, 10, step=1,
                                     key="children_input", on_change=_sync_family)
        _n_ch    = int(st.session_state.get("children_input", 0))
        _married = bool(st.session_state.get("married_input", True))
        min_fam  = _n_ch + (2 if _married else 1)
        family_size = r1c4.number_input(
            "Размер семьи", min_value=min_fam, max_value=15,
            key="family_size_input",
            help=f"Авто-минимум: {'дети+супруг+вы' if _married else 'дети+вы'} = {min_fam}",
        )

        # Строка 2: чекбоксы — все на одном уровне
        r2c1, r2c2, r2c3 = st.columns(3)
        married    = r2c1.checkbox("Женат / замужем",  key="married_input",  on_change=_sync_family)
        own_car    = r2c2.checkbox("Есть автомобиль",  key="car_input")
        own_realty = r2c3.checkbox("Есть недвижимость", key="realty_input")

    # ── Финансовые данные ─────────────────────────────────────────────────────
    with st.expander("▼ Финансовые данные", expanded=True):
        fc1, fc2, fc3, fc4 = st.columns(4)
        income  = fc1.number_input("Доход (сом/мес)", min_value=0, step=5_000, key="income_input")
        credit  = fc2.number_input("Сумма кредита (сом)", min_value=0, step=10_000,
                                   key="credit_input", on_change=_sync_annuity)
        term    = fc3.number_input("Срок (мес.)", min_value=6, max_value=360, step=6,
                                   key="term_input", on_change=_sync_annuity,
                                   help="Влияет на авто-расчёт платежа")
        annuity = fc4.number_input("Платёж (сом/мес)", min_value=0, step=1_000,
                                   key="annuity_input",
                                   help="Авто-рассчитывается по кредиту и сроку")

        ec1, ec2 = st.columns(2)
        years_employed = ec1.number_input("Трудовой стаж (лет)", 0, 50, step=1, key="emp_input")
        # checkbox не выравнивается с number_input — добавляем отступ
        ec2.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
        is_unemployed  = ec2.checkbox("Безработный", key="unemp_input")

    # ── Quick stats ───────────────────────────────────────────────────────────
    show_quick_stats(float(income), float(credit), float(annuity), int(term))

    # ── Кредитная история ─────────────────────────────────────────────────────
    with st.expander("▶ Кредитная история (необязательно)", expanded=False):
        st.caption("0 = плохая история, 1 = отличная. Снимите галочку, если данных нет.")
        sc1, sc2, sc3 = st.columns(3)

        with sc1:
            use_ext1 = st.checkbox("Данные бюро 1", key="use_ext1_input")
            ext1 = st.slider("Скоринг бюро 1", 0.0, 1.0, step=0.01,
                             key="ext1_input", disabled=not use_ext1)

        with sc2:
            use_ext2 = st.checkbox("Данные бюро 2", key="use_ext2_input")
            ext2 = st.slider("Скоринг бюро 2", 0.0, 1.0, step=0.01,
                             key="ext2_input", disabled=not use_ext2)

        with sc3:
            use_ext3 = st.checkbox("Данные бюро 3", key="use_ext3_input")
            ext3 = st.slider("Скоринг бюро 3", 0.0, 1.0, step=0.01,
                             key="ext3_input", disabled=not use_ext3)

    btn_col, clear_col = st.columns([5, 1])
    submitted = btn_col.button("🔍 Оценить заёмщика", type="primary", use_container_width=True)
    if clear_col.button("🗑️ Очистить", use_container_width=True,
                        disabled="tab1_result" not in st.session_state):
        del st.session_state["tab1_result"]

    # ── Вычисление при нажатии — сохраняем в session_state ───────────────────
    if submitted:
        form_data = dict(
            age=int(age), gender=gender, married=bool(married),
            children=int(children), family_size=int(family_size),
            own_car=bool(own_car), own_realty=bool(own_realty),
            income=float(income), credit=float(credit), annuity=float(annuity),
            years_employed=float(years_employed), is_unemployed=bool(is_unemployed),
            ext_source_1=float(ext1) if use_ext1 else None,
            ext_source_2=float(ext2) if use_ext2 else None,
            ext_source_3=float(ext3) if use_ext3 else None,
        )

        X_input = prepare_single_input(form_data, medians, feature_cols)
        prob = float(predict_proba_safe(model, X_input)[0])
        model_decision = "ОТКАЗАТЬ" if prob >= threshold else "ОДОБРИТЬ"

        rule_violations = check_business_rules(income, credit, annuity)
        if rule_violations:
            for v in rule_violations:
                logger.warning("БИЗНЕС-ПРАВИЛО НАРУШЕНО: %s", v)
            decision = "ОТКАЗАТЬ"
        else:
            decision = model_decision

        logger.info(
            "Оценка | модель=%s | prob=%.4f | порог=%.4f | ML=%s | итог=%s | "
            "DTI=%.1f%% | кредит/доход=%.1f× | доход=%.0f | кредит=%.0f | платёж=%.0f",
            selected_model_name.upper(), prob, threshold, model_decision, decision,
            annuity / max(income, 1) * 100, credit / max(income, 1),
            income, credit, annuity,
        )

        shap_vals, expected_value = compute_shap_values(
            model, selected_model_name, X_input, X_background
        )
        if shap_vals is None and selected_model_name == "ebm":
            ebm_scores, ebm_names = get_ebm_local_importance(model, X_input)
            if ebm_scores is not None:
                name_to_idx = {n: i for i, n in enumerate(feature_cols)}
                shap_vals = np.zeros(len(feature_cols))
                for n, s in zip(ebm_names, ebm_scores):
                    if n in name_to_idx:
                        shap_vals[name_to_idx[n]] = s
                expected_value = 0.0

        st.session_state["tab1_result"] = dict(
            prob=prob, rule_violations=rule_violations,
            shap_vals=shap_vals, expected_value=expected_value,
            X_input=X_input, income=income, credit=credit, annuity=annuity,
            model_name=selected_model_name,
        )

    # ── Рендер результата из session_state ───────────────────────────────────
    if "tab1_result" in st.session_state:
        r = st.session_state["tab1_result"]
        prob            = r["prob"]
        rule_violations = r["rule_violations"]
        shap_vals       = r["shap_vals"]
        expected_value  = r["expected_value"]
        X_input         = r["X_input"]
        _income         = r["income"]
        _credit         = r["credit"]
        _annuity        = r["annuity"]

        # Зона рассчитывается динамически (не хранится в state)
        zone_code, zone_label, zone_color, zone_decision, zone_icon = get_zone(prob, threshold)

        st.markdown("---")

        # Block A: Decision banner — по зонам риска

        # Бизнес-правила перекрывают зону (всегда отказ)
        if rule_violations:
            zone_code     = "very_high"
            zone_label    = "Очень высокий"
            zone_color    = "#dc3545"
            zone_decision = "АВТО-ОТКАЗ"
            zone_icon     = "❌"

        banner_text = f"## {zone_icon}  {zone_decision}"
        if zone_code in ("very_low", "low", "medium"):
            st.success(banner_text)
        elif zone_code == "borderline":
            st.warning(banner_text)
        else:
            st.error(banner_text)

        st.caption(
            f"Зона риска: **{zone_label}** · "
            f"Вероятность дефолта: **{prob:.1%}** · "
            f"Порог F₂ (β=2): **{threshold:.4f}**"
        )

        if rule_violations:
            st.warning(
                "**Автоматический отказ по бизнес-правилам** "
                "(независимо от вероятности модели)\n\n"
                + "\n".join(f"- {v}" for v in rule_violations)
            )

        # Metrics row
        if show_mode_specialist and show_mode_analyst:
            m1, m2, m3, m4, m5, m6 = st.columns(6)
            m1.metric("Зона риска",          zone_label)
            m2.metric("Долговая нагрузка",   f"{_annuity/max(_income,1):.1%}")
            m3.metric("Кредит к доходу",     f"{_credit/max(_income,1):.1f}×")
            m4.metric("Вероятность дефолта", f"{prob:.1%}")
            m5.metric("Порог (F₂, β=2)",     f"{threshold:.3f}")
            m6.metric("Отступ от порога",    f"{(prob - threshold)*100:+.1f} п.п.")
        elif show_mode_specialist:
            m1, m2, m3 = st.columns(3)
            m1.metric("Зона риска",        zone_label)
            m2.metric("Долговая нагрузка", f"{_annuity/max(_income,1):.1%}")
            m3.metric("Кредит к доходу",   f"{_credit/max(_income,1):.1f}×")
        else:
            m1, m2, m3 = st.columns(3)
            m1.metric("Вероятность дефолта", f"{prob:.1%}")
            m2.metric("Порог (F₂, β=2)",     f"{threshold:.3f}")
            m3.metric("Отступ от порога",    f"{(prob - threshold)*100:+.1f} п.п.")

        # Block B: Gauge
        show_gauge(prob, threshold)

        # Block C: Factors
        st.markdown("---")
        st.markdown("### 📊 Почему такое решение?")
        if shap_vals is not None:
            show_factors(shap_vals, feature_cols, X_input, display_mode)
        else:
            st.info("Объяснение факторов недоступно для данной модели.")

        # Block D: SHAP waterfall (analyst only)
        if show_mode_analyst and shap_vals is not None and expected_value is not None:
            with st.expander("📈 Детальный SHAP-анализ (для аналитика)", expanded=(display_mode == "Только для аналитика")):
                show_waterfall(shap_vals, expected_value, feature_cols, X_input)
                st.caption(
                    f"**Техническая информация:** "
                    f"Модель: {r['model_name'].upper()} | "
                    f"Порог: {threshold:.4f} (F₂, β=2) | "
                    f"Вероятность: {prob:.4f} | "
                    f"Решение: {'Дефолт' if prob >= threshold else 'Не дефолт'}"
                )

        # Block E: Recommendation
        st.markdown("---")
        show_recommendation(prob, threshold)


# ═══════════════════════════════════════════════════════════════════════════════
#  TAB 2 — Batch upload
# ═══════════════════════════════════════════════════════════════════════════════
with tab2:
    st.header("Пакетная оценка заявок")

    st.info(
        "📋 **Пакетная оценка заявок**\n\n"
        "Загрузите файл CSV или Excel со списком заявок. "
        "Система автоматически оценит риск по каждой заявке.\n\n"
        f"Обязательные колонки: **{', '.join(REQUIRED_BATCH_COLS)}**. "
        "Остальные данные будут заполнены средними значениями по базе."
    )

    up_col, dl_col = st.columns([3, 1])
    with up_col:
        uploaded = st.file_uploader(
            "Выберите файл",
            type=["csv", "xlsx", "xls"],
            help="Максимальный размер: 200 МБ",
        )
    with dl_col:
        st.markdown("<br>", unsafe_allow_html=True)
        st.download_button(
            "📥 Скачать шаблон",
            data=make_template_csv(),
            file_name="template_scoring.csv",
            mime="text/csv",
        )

    if uploaded is not None:
        # Load file
        try:
            if uploaded.name.endswith((".xlsx", ".xls")):
                raw_df = pd.read_excel(uploaded)
            else:
                raw_df = pd.read_csv(uploaded)
        except Exception as e:
            st.error(f"❌ Не удалось прочитать файл. Проверьте формат (CSV или Excel).\n\n{e}")
            st.stop()

        if raw_df.empty:
            st.error("❌ Файл пустой или не содержит данных.")
            st.stop()

        st.success(f"✅ Файл загружен: **{uploaded.name}** · {len(raw_df):,} заявок · {len(raw_df.columns)} колонок")

        # Preview
        st.dataframe(raw_df.head(5), use_container_width=True)

        missing_required = [c for c in REQUIRED_BATCH_COLS if c not in raw_df.columns]
        if missing_required:
            st.warning(f"⚠️ Отсутствующие обязательные колонки: **{', '.join(missing_required)}**. Будут заполнены средними.")

        # Run scoring
        if st.button("▶ Запустить оценку", type="primary", use_container_width=True):
            progress = st.progress(0, text="Подготовка данных…")

            try:
                X_batch = prepare_batch_input(raw_df, medians, feature_cols)
                progress.progress(40, text="Предсказание вероятностей…")

                probs = predict_proba_safe(model, X_batch)
                progress.progress(80, text="Формирование результатов…")

                result_df = raw_df.copy()
                result_df["default_prob"] = probs
                result_df["default_prob_pct"] = (probs * 100).round(1)
                result_df["risk_level"] = [risk_level_label(p, threshold) for p in probs]
                result_df["decision"] = ["Отказать" if p >= threshold else "Одобрить" for p in probs]
                if ID_COL not in result_df.columns:
                    result_df.insert(0, ID_COL, range(1, len(result_df) + 1))

                # FIX 5: применяем бизнес-правила к каждой строке батча
                if "AMT_INCOME_TOTAL" in raw_df.columns and "AMT_ANNUITY" in raw_df.columns:
                    _inc = raw_df["AMT_INCOME_TOTAL"].clip(lower=1)
                    _ann = raw_df["AMT_ANNUITY"]
                    _cred_col = (raw_df["AMT_CREDIT"] if "AMT_CREDIT" in raw_df.columns
                                 else raw_df.get("AMT_GOODS_PRICE",
                                                  pd.Series(medians.get("AMT_GOODS_PRICE", 0),
                                                             index=raw_df.index)))
                    _dti = _ann / _inc
                    _cti = _cred_col / _inc
                    _hard = (_dti >= 0.80) | (_cti > 240)
                    if _hard.any():
                        n_hard = int(_hard.sum())
                        result_df.loc[_hard.values, "decision"]   = "Отказать"
                        result_df.loc[_hard.values, "risk_level"] = "Очень высокий"
                        logger.warning("Батч: %d строк принудительно отклонено по бизнес-правилам", n_hard)

                progress.progress(100, text="Готово!")

            except Exception as e:
                logger.error("Ошибка батч-обработки: %s", e, exc_info=True)
                st.error(f"❌ Ошибка при обработке: {e}")
                st.stop()

            # Предупреждение о бизнес-правилах (если были override'ы)
            if "AMT_INCOME_TOTAL" in raw_df.columns and "AMT_ANNUITY" in raw_df.columns:
                _inc = raw_df["AMT_INCOME_TOTAL"].clip(lower=1)
                _ann = raw_df["AMT_ANNUITY"]
                _cred_col = (raw_df["AMT_CREDIT"] if "AMT_CREDIT" in raw_df.columns
                             else raw_df.get("AMT_GOODS_PRICE",
                                              pd.Series(medians.get("AMT_GOODS_PRICE", 0),
                                                         index=raw_df.index)))
                _hard = ((_ann / _inc) >= 0.80) | ((_cred_col / _inc) > 240)
                if _hard.any():
                    st.warning(
                        f"⚠️ **{int(_hard.sum())} заявок** принудительно отклонено по бизнес-правилам "
                        f"(DTI ≥ 80% или кредит > 240× дохода) — независимо от вероятности модели."
                    )

            # ── Summary panel ────────────────────────────────────────────────
            st.markdown("---")
            n_total    = len(result_df)
            n_approved = (result_df["decision"] == "Одобрить").sum()
            n_rejected = (result_df["decision"] == "Отказать").sum()
            mean_prob  = probs.mean()

            logger.info(
                "Батч | файл=%s | модель=%s | строк=%d | одобрено=%d (%.1f%%) | "
                "отказано=%d (%.1f%%) | средний_риск=%.1f%% | порог=%.4f",
                uploaded.name, selected_model_name.upper(),
                n_total, n_approved, n_approved / n_total * 100,
                n_rejected, n_rejected / n_total * 100,
                mean_prob * 100, threshold,
            )

            sm1, sm2, sm3, sm4 = st.columns(4)
            sm1.metric("Всего заявок",   f"{n_total:,}")
            sm2.metric("Одобрено ✅",     f"{n_approved:,} ({n_approved/n_total:.1%})")
            sm3.metric("Отказано ❌",     f"{n_rejected:,} ({n_rejected/n_total:.1%})")
            sm4.metric("Средний риск",    f"{mean_prob:.1%}")

            # Analyst metrics if TARGET present
            if TARGET_COL in raw_df.columns and show_mode_analyst:
                y_true_batch = raw_df[TARGET_COL].values
                y_pred_batch = (probs >= threshold).astype(int)
                am1, am2, am3, am4 = st.columns(4)
                am1.metric("ROC-AUC",   f"{roc_auc_score(y_true_batch, probs):.4f}",  help="Чем ближе к 1 — тем лучше.")
                am2.metric("F₂-score",  f"{fbeta_score(y_true_batch, y_pred_batch, beta=2, zero_division=0):.4f}", help="F₂-мера (β=2): recall важнее precision.")
                am3.metric("Recall",    f"{recall_score(y_true_batch, y_pred_batch, zero_division=0):.1%}", help="Доля реальных дефолтов, выявленных моделью.")
                am4.metric("Precision", f"{precision_score(y_true_batch, y_pred_batch, zero_division=0):.1%}", help="Точность отказов.")

            # ── Risk histogram ───────────────────────────────────────────────
            fig_hist = px.histogram(
                result_df,
                x="default_prob_pct",
                color="decision",
                color_discrete_map={"Одобрить": "#28a745", "Отказать": "#dc3545"},
                nbins=25,
                labels={"default_prob_pct": "Риск невозврата (%)", "count": "Количество заявок"},
                title="Распределение заявок по уровню риска",
            )
            fig_hist.add_vline(
                x=threshold * 100,
                line_dash="dash",
                line_color="black",
                annotation_text=f"Порог {threshold:.1%}",
                annotation_position="top right",
            )
            if show_mode_analyst:
                fig_hist.add_annotation(
                    x=threshold * 100, y=0,
                    text=f"threshold={threshold:.4f}<br>(F₂, β=2)",
                    showarrow=True, arrowhead=2, bgcolor="white", bordercolor="black",
                )
            st.plotly_chart(fig_hist, use_container_width=True)

            # ── Results table ────────────────────────────────────────────────
            st.markdown("### Результаты оценки")

            if show_mode_specialist and not show_mode_analyst:
                display_rus = {
                    ID_COL:             "ID заявки",
                    "AMT_INCOME_TOTAL": "Годовой доход",
                    "AMT_GOODS_PRICE":  "Сумма кредита",
                    "AMT_ANNUITY":      "Платёж/мес",
                    "default_prob_pct": "Риск (%)",
                    "risk_level":       "Уровень риска",
                    "decision":         "Решение",
                }
                cols_to_show = [c for c in display_rus if c in result_df.columns]
                show_df = result_df[cols_to_show].rename(columns=display_rus)
            else:
                cols_to_show = [ID_COL, "AMT_INCOME_TOTAL", "AMT_GOODS_PRICE",
                                "AMT_ANNUITY", "default_prob", "default_prob_pct",
                                "risk_level", "decision"]
                cols_to_show = [c for c in cols_to_show if c in result_df.columns]
                show_df = result_df[cols_to_show]

            def _color_rows(row):
                val = row["Решение"] if "Решение" in row.index else (row["decision"] if "decision" in row.index else "")
                color = "#fff0f0" if val == "Отказать" else "#f0fff0"
                return [f"background-color: {color}"] * len(row)

            styled = show_df.style.apply(_color_rows, axis=1)
            st.dataframe(styled, use_container_width=True, height=400)

            # ── Analytics expander (analyst) ─────────────────────────────────
            if show_mode_analyst:
                with st.expander("📈 Аналитика по портфелю (для аналитика)", expanded=False):
                    an1, an2 = st.columns(2)

                    with an1:
                        st.subheader("⚠️ Топ-10 рискованных заявок")
                        top10_cols = [ID_COL, "default_prob_pct", "risk_level", "decision"]
                        top10_cols = [c for c in top10_cols if c in result_df.columns]
                        st.dataframe(
                            result_df.nlargest(10, "default_prob")[top10_cols],
                            hide_index=True, use_container_width=True,
                        )

                    with an2:
                        risk_counts = result_df["risk_level"].value_counts()
                        # FIX 6: color_map строится из ZONE_DATA — названия всегда совпадают
                        color_map = {label: color for _, label, color, _, _ in ZONE_DATA}
                        fig_pie = px.pie(
                            values=risk_counts.values,
                            names=risk_counts.index,
                            color=risk_counts.index,
                            color_discrete_map=color_map,
                            title="Структура портфеля по уровням риска",
                        )
                        st.plotly_chart(fig_pie, use_container_width=True)

                    if TARGET_COL in raw_df.columns:
                        st.markdown("---")
                        st.subheader("📊 Качество модели на загруженных данных")
                        bm = business_metric(y_true_batch, y_pred_batch)
                        st.info(
                            f"💰 **Бизнес-метрика:**  \n"
                            f"Пропущено дефолтов (FN): **{bm['fn']}** — потенциальные убытки банка  \n"
                            f"Ложных отказов (FP): **{bm['fp']}** — потерянные хорошие клиенты  \n"
                            f"Суммарные издержки (FN×1 + FP×5): **{bm['total_cost']:.0f}** усл. единиц"
                        )

            # ── Download buttons ──────────────────────────────────────────────
            st.markdown("---")
            dl1, dl2 = st.columns(2)
            ts = datetime.now().strftime("%Y%m%d_%H%M")

            excel_bytes = generate_excel_report(result_df, threshold, selected_model_name)
            dl1.download_button(
                "⬇️ Скачать результаты (Excel)",
                data=excel_bytes,
                file_name=f"scoring_results_{ts}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
            )

            csv_bytes = result_df.to_csv(index=False, encoding="utf-8-sig").encode("utf-8-sig")
            dl2.download_button(
                "⬇️ Скачать результаты (CSV)",
                data=csv_bytes,
                file_name=f"scoring_results_{ts}.csv",
                mime="text/csv",
            )
