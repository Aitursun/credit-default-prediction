from __future__ import annotations

import io
import json
import logging
import warnings
from datetime import datetime
from pathlib import Path

import joblib
import numpy as np
import pandas as pd
import shap

warnings.filterwarnings("ignore")

ROOT = Path(__file__).parent.parent
import sys
sys.path.insert(0, str(ROOT))

from src.config import ID_COL, MODELS_DIR, PROC_DATA_DIR, TARGET_COL
from src.evaluation import find_optimal_threshold

logger = logging.getLogger("credit_scoring")

# ─── Словарь объяснений признаков ────────────────────────────────────────────

FEATURE_EXPLANATIONS: dict[str, tuple[str, str, str, str]] = {
    "EXT_SOURCE_MEAN":     ("Средний скоринг кредитных бюро",  "Низкий — плохая история погашений",        "Высокий — хорошая история погашений",       "0–1"),
    "EXT_SOURCE_2":        ("Скоринг бюро 2",                  "Низкий — плохая история",                  "Высокий — хорошая история",                 "0–1"),
    "EXT_SOURCE_3":        ("Скоринг бюро 3",                  "Низкий — плохая история",                  "Высокий — хорошая история",                 "0–1"),
    "EXT_SOURCE_1":        ("Скоринг бюро 1",                  "Низкий — плохая история",                  "Высокий — хорошая история",                 "0–1"),
    "EXT_SOURCE_MIN":      ("Минимальный скоринг бюро",        "Один из скорингов очень низкий",            "Все скоринги в норме",                      "0–1"),
    "EXT_SOURCE_MAX":      ("Максимальный скоринг бюро",       "Даже лучший скоринг низкий",               "Лучший скоринг высокий",                    "0–1"),
    "EXT_SOURCE_STD":      ("Разброс скорингов бюро",          "Большой разброс — нестабильность",         "Скоринги согласованны",                     ""),
    "ANNUITY_INCOME_RATIO":("Долговая нагрузка (платёж/доход)","Превышает норму 30% — высокий риск",       "В пределах нормы",                          "%"),
    "CREDIT_INCOME_RATIO": ("Отношение кредита к доходу",      "Кредит значительно превышает доход",       "Кредит соразмерен доходу",                  "×"),
    "ANNUITY_CREDIT_RATIO":("Доля платежа от кредита",         "Высокая доля — короткий срок погашения",   "Умеренная доля платежа",                    ""),
    "CREDIT_GOODS_RATIO":  ("Кредит к стоимости товара",       "Кредит превышает стоимость товара",        "Кредит соответствует стоимости",            ""),
    "AMT_INCOME_TOTAL":    ("Месячный доход",                  "Низкий доход повышает риск",               "Высокий доход снижает риск",                "сом"),
    "AMT_ANNUITY":         ("Ежемесячный платёж",              "Высокий платёж",                           "Умеренный платёж",                          "сом"),
    "AMT_GOODS_PRICE":     ("Сумма кредита",                   "Очень крупный кредит",                     "Умеренная сумма",                           "сом"),
    "DAYS_BIRTH":          ("Возраст заёмщика",                "Молодой возраст — статистически выше риск","Зрелый возраст — ниже риск",                "лет"),
    "YEARS_EMPLOYED":      ("Трудовой стаж",                   "Короткий стаж — нестабильная занятость",   "Длительный стаж — стабильная занятость",    "лет"),
    "IS_UNEMPLOYED":       ("Статус занятости",                "Безработный — значительно повышает риск",  "Трудоустроен",                              ""),
    "FLAG_OWN_REALTY_Y":   ("Наличие недвижимости",            "Нет недвижимости",                        "Есть недвижимость — снижает риск",          ""),
    "FLAG_OWN_CAR_Y":      ("Наличие автомобиля",              "Нет автомобиля",                          "Есть автомобиль — снижает риск",            ""),
    "CNT_CHILDREN":        ("Количество детей",                "Много детей — высокая финансовая нагрузка","Детей нет или мало",                        ""),
    "CNT_FAM_MEMBERS":     ("Размер семьи",                    "Большая семья — высокая нагрузка",         "Небольшая семья",                           ""),
    "INCOME_PER_PERSON":   ("Доход на члена семьи",            "Низкий доход на человека",                 "Достаточный доход на члена семьи",          "сом"),
    "CODE_GENDER_M":       ("Пол заёмщика",                    "Мужской",                                 "Женский",                                   ""),
    "AMT_REQ_CREDIT_BUREAU_TOTAL": ("Запросы в кредитное бюро","Много запросов — активно ищет кредиты",   "Запросов мало",                             "шт."),
    "DOCUMENT_COUNT":      ("Количество документов",           "Мало документов",                          "Полный пакет документов",                   "шт."),
    "BUREAU_LOAN_COUNT":   ("Кредитов в истории",              "Много кредитов",                           "Умеренное количество",                      "шт."),
    "BUREAU_ACTIVE_COUNT": ("Активных кредитов сейчас",        "Много активных кредитов",                  "Мало активных кредитов",                    "шт."),
    "BUREAU_ACTIVE_OVERDUE_SUM":("Сумма просрочек по активным","Есть просрочки — серьёзный риск",          "Просрочек нет",                             "сом"),
    "BB_STATUS_DPD_COUNT_SUM":  ("Случаев просрочки в истории","Много просрочек в прошлом",                "Просрочек нет",                             ""),
    "INS_DPD_MEAN":        ("Среднее дней просрочки",          "Высокая просрочка по рассрочкам",          "Платит вовремя",                            "дней"),
    "INS_OVERDUE_RATIO":   ("Доля просроченных платежей",      "Высокая доля просрочек",                   "Платит стабильно",                          "%"),
    "PREV_REFUSED_RATIO":  ("Доля отказов в прошлом",          "Часто отказывали — плохая история",        "Преимущественно одобряли",                  "%"),
    "POS_SK_DPD_DEF_MEAN": ("Дефолтные просрочки (POS)",       "Есть дефолтные просрочки",                 "Дефолтных просрочек нет",                   ""),
    "CC_SK_DPD_DEF_MEAN":  ("Дефолтные просрочки (карта)",     "Есть дефолтные просрочки по карте",        "Дефолтных просрочек нет",                   ""),
    "CC_UTILIZATION_MEAN": ("Утилизация кредитной карты",      "Высокая утилизация карты",                 "Умеренная утилизация",                      "%"),
    "REGION_RATING_CLIENT_W_CITY": ("Рейтинг региона",         "Неблагоприятный регион проживания",        "Благоприятный регион",                      ""),
    "DAYS_REGISTRATION_AGE_RATIO": ("Регистрация / возраст",   "Поздняя регистрация относительно возраста","Ранняя регистрация",                        ""),
    "DAYS_ID_PUBLISH_AGE_RATIO":   ("Смена документов / возраст","Поздняя смена документов",              "Документы в порядке",                       ""),
    "PHONE_TO_BIRTH_RATIO":        ("Смена телефона / возраст","Частая смена телефона",                    "Стабильный контакт",                        ""),
    "OWN_CAR_AGE":         ("Возраст автомобиля",              "Старый автомобиль",                        "Новый автомобиль",                          "лет"),
    "DEF_30_CNT_SOCIAL_CIRCLE":    ("Дефолты в окружении",     "Много дефолтов у знакомых",                "В окружении нет дефолтов",                  ""),
    "NAME_EDUCATION_TYPE_Higher education":             ("Образование: высшее",       "", "", ""),
    "NAME_EDUCATION_TYPE_Secondary / secondary special":("Образование: среднее",      "", "", ""),
    "NAME_EDUCATION_TYPE_Incomplete higher":            ("Образование: неполное высшее","","",""),
    "NAME_EDUCATION_TYPE_Lower secondary":              ("Образование: неполное среднее","","",""),
}

# ─── Зоны риска (Базель II IRB, EBA GL/2020/06) ──────────────────────────────

ZONE_DATA = [
    # (код,         label_ru,          цвет,     label_решения,         иконка)
    ("very_low",   "Очень низкий",  "#1a7a3c", "АВТО-ОДОБРЕНИЕ",      "✅"),
    ("low",        "Низкий",        "#28a745", "ОДОБРИТЬ",             "✅"),
    ("medium",     "Средний",       "#6db33f", "ОДОБРИТЬ С УСЛОВИЕМ",  "✅"),
    ("borderline", "Пограничный",   "#e6a817", "РУЧНАЯ ПРОВЕРКА",      "⚠️"),
    ("high",       "Высокий",       "#e07320", "УСЛОВНЫЙ ОТКАЗ",       "⛔"),
    ("very_high",  "Очень высокий", "#dc3545", "АВТО-ОТКАЗ",           "❌"),
]

# ─── Маппинг образования ─────────────────────────────────────────────────────

EDU_OPTIONS = [
    "Среднее / среднее специальное",
    "Высшее образование",
    "Неполное высшее",
    "Неполное среднее",
]

EDU_MAP = {
    "Высшее образование":            "NAME_EDUCATION_TYPE_Higher education",
    "Неполное высшее":               "NAME_EDUCATION_TYPE_Incomplete higher",
    "Неполное среднее":              "NAME_EDUCATION_TYPE_Lower secondary",
    "Среднее / среднее специальное": "NAME_EDUCATION_TYPE_Secondary / secondary special",
}

# ─── Пороги ──────────────────────────────────────────────────────────────────

THRESHOLDS_PATH = PROC_DATA_DIR / "thresholds.json"


def load_thresholds() -> dict:
    if THRESHOLDS_PATH.exists():
        return json.loads(THRESHOLDS_PATH.read_text())
    return {}


def compute_threshold_holdout(model_name: str, model, beta: float = 2.0) -> tuple[float, float]:
    """Fallback: вычисляет порог на 30% holdout если JSON не найден."""
    from sklearn.model_selection import train_test_split
    from src.training import load_data
    X, y = load_data()
    y_arr = y.to_numpy() if hasattr(y, "to_numpy") else y
    _, X_val, _, y_val = train_test_split(X, y_arr, test_size=0.3, stratify=y_arr, random_state=42)
    probs = predict_proba_safe(model, X_val)
    return find_optimal_threshold(y_val, probs, beta=beta)


# ─── Загрузка данных ─────────────────────────────────────────────────────────

def load_model_from_disk(name: str):
    path = MODELS_DIR / f"{name}.joblib"
    if not path.exists():
        raise FileNotFoundError(f"Модель не найдена: {path}. Запустите ноутбук 04.")
    m = joblib.load(path)
    logger.info("Модель загружена: %s (%s) scaler=%s", name.upper(), type(m).__name__, hasattr(m, "_scaler"))
    return m


def load_train_stats_from_disk() -> tuple[pd.Series, list[str], pd.DataFrame]:
    """Возвращает (медианы, feature_cols, X_background)."""
    from src.training import load_data
    X, _ = load_data()
    feature_cols = list(X.columns)
    medians = X.median()
    shap_cache = joblib.load(MODELS_DIR / "lgbm_shap.joblib")
    X_background: pd.DataFrame = shap_cache["X_sample"]
    return medians, feature_cols, X_background


# ─── Предсказание ────────────────────────────────────────────────────────────

def predict_proba_safe(model, X: pd.DataFrame) -> np.ndarray:
    """Применяет _scaler если модель была обучена с ним, возвращает P(дефолт)."""
    X_in = X.to_numpy(dtype=np.float32)
    if hasattr(model, "_scaler"):
        X_in = model._scaler.transform(X_in)
    return model.predict_proba(X_in)[:, 1]


def scale_if_needed(model, X) -> np.ndarray:
    arr = X.to_numpy(dtype=np.float32) if isinstance(X, pd.DataFrame) else X.astype(np.float32)
    if hasattr(model, "_scaler"):
        arr = model._scaler.transform(arr)
    return arr


def build_shap_explainer(model, X_background: pd.DataFrame, model_name: str):
    """Создаёт SHAP explainer под тип модели."""
    try:
        bg = scale_if_needed(model, X_background[:500])
        if model_name == "logreg":
            return shap.LinearExplainer(model, bg)
        elif model_name == "ebm":
            return None
        else:
            return shap.TreeExplainer(
                model, scale_if_needed(model, X_background[:200]),
                feature_perturbation="interventional",
            )
    except Exception:
        try:
            return shap.TreeExplainer(model)
        except Exception:
            return None


def compute_shap_values(model, model_name: str, X_input: pd.DataFrame, explainer) -> tuple:
    """Возвращает (shap_values_1d, expected_value) или (None, None)."""
    if explainer is None:
        return None, None
    try:
        X_in = scale_if_needed(model, X_input)
        shap_exp = explainer(X_in)
        vals = shap_exp.values
        if vals.ndim == 3:
            vals = vals[:, :, 1]
        ev = float(shap_exp.base_values[0]) if vals.ndim == 2 else float(shap_exp.base_values)
        if isinstance(ev, np.ndarray):
            ev = float(ev.flat[0])
        return vals[0], ev
    except Exception:
        return None, None


def get_ebm_local_importance(model, X_input: pd.DataFrame) -> tuple:
    """Извлекает локальные вклады из EBM explain_local."""
    try:
        local_exp = model.explain_local(X_input, name="local")
        data = local_exp.data(0)
        return np.array(data["scores"]), list(data["names"])
    except Exception:
        return None, None


# ─── Подготовка признаков ─────────────────────────────────────────────────────

def prepare_single_input(form: dict, medians: pd.Series, feature_cols: list[str]) -> pd.DataFrame:
    """Преобразует данные формы в вектор из 209 признаков."""
    row = medians.copy()

    income  = max(form["income"], 1)
    annuity = form["annuity"]
    credit  = max(form["credit"], 1)

    row["DAYS_BIRTH"]       = -form["age"] * 365
    row["AMT_INCOME_TOTAL"] = income
    row["AMT_ANNUITY"]      = annuity
    row["AMT_GOODS_PRICE"]  = credit
    if "AMT_CREDIT" in row.index:
        row["AMT_CREDIT"]   = credit
    row["CNT_CHILDREN"]     = form["children"]
    row["CNT_FAM_MEMBERS"]  = form["family_size"]
    row["CODE_GENDER_M"]    = 1 if form["gender"] == "Мужской" else 0
    row["FLAG_OWN_CAR_Y"]   = int(form["own_car"])
    row["FLAG_OWN_REALTY_Y"]= int(form["own_realty"])

    if form["is_unemployed"]:
        row["IS_UNEMPLOYED"]  = 1
        row["YEARS_EMPLOYED"] = 0.0
    else:
        row["IS_UNEMPLOYED"]  = 0
        row["YEARS_EMPLOYED"] = form["years_employed"]

    for i in range(1, 4):
        val = form.get(f"ext_source_{i}")
        col, miss = f"EXT_SOURCE_{i}", f"EXT_SOURCE_{i}_MISSING"
        if val is not None:
            row[col] = val
            if miss in row.index:
                row[miss] = 0

    row["CREDIT_INCOME_RATIO"]  = credit / income
    row["ANNUITY_INCOME_RATIO"] = annuity / income
    row["CREDIT_GOODS_RATIO"]   = 1.0
    row["ANNUITY_CREDIT_RATIO"] = annuity / credit

    ext_vals = [row[f"EXT_SOURCE_{i}"] for i in range(1, 4)
                if not np.isnan(row.get(f"EXT_SOURCE_{i}", np.nan))]
    if ext_vals:
        row["EXT_SOURCE_MEAN"] = float(np.mean(ext_vals))
        row["EXT_SOURCE_MIN"]  = float(np.min(ext_vals))
        row["EXT_SOURCE_MAX"]  = float(np.max(ext_vals))
        row["EXT_SOURCE_STD"]  = float(np.std(ext_vals)) if len(ext_vals) > 1 else 0.0

    row["INCOME_PER_PERSON"] = income / max(form["family_size"], 1)

    # Семейный статус
    married = form.get("married", True)
    for col in ["NAME_FAMILY_STATUS_Married", "NAME_FAMILY_STATUS_Single / not married",
                "NAME_FAMILY_STATUS_Civil marriage", "NAME_FAMILY_STATUS_Separated",
                "NAME_FAMILY_STATUS_Widow"]:
        if col in row.index:
            row[col] = 0
    if "NAME_FAMILY_STATUS_Married" in row.index:
        row["NAME_FAMILY_STATUS_Married"] = 1 if married else 0
    if "NAME_FAMILY_STATUS_Single / not married" in row.index:
        row["NAME_FAMILY_STATUS_Single / not married"] = 0 if married else 1

    # Образование
    for col in EDU_MAP.values():
        if col in row.index:
            row[col] = 0
    target = EDU_MAP.get(form.get("education", "Среднее / среднее специальное"))
    if target and target in row.index:
        row[target] = 1

    return pd.DataFrame([row[feature_cols]], columns=feature_cols)


def prepare_batch_input(uploaded_df: pd.DataFrame, medians: pd.Series,
                        feature_cols: list[str]) -> pd.DataFrame:
    """Преобразует загруженный CSV в матрицу признаков."""
    out = pd.DataFrame(
        np.tile(medians[feature_cols].values, (len(uploaded_df), 1)),
        columns=feature_cols,
    )
    for col in feature_cols:
        if col in uploaded_df.columns:
            out[col] = uploaded_df[col].values

    if "AGE_YEARS" in uploaded_df.columns and "DAYS_BIRTH" in feature_cols:
        out["DAYS_BIRTH"] = -uploaded_df["AGE_YEARS"].values * 365

    if "AMT_CREDIT" in uploaded_df.columns:
        credit  = uploaded_df["AMT_CREDIT"].clip(lower=1).values
        income  = uploaded_df.get("AMT_INCOME_TOTAL",
                  pd.Series([medians["AMT_INCOME_TOTAL"]] * len(uploaded_df))).clip(lower=1).values
        annuity = uploaded_df.get("AMT_ANNUITY",
                  pd.Series([medians["AMT_ANNUITY"]] * len(uploaded_df))).values
        out["AMT_GOODS_PRICE"]      = credit
        out["CREDIT_INCOME_RATIO"]  = credit / income
        out["ANNUITY_INCOME_RATIO"] = annuity / income
        out["CREDIT_GOODS_RATIO"]   = 1.0
        out["ANNUITY_CREDIT_RATIO"] = annuity / credit

    ext_present = [f"EXT_SOURCE_{i}" in uploaded_df.columns for i in range(1, 4)]
    if any(ext_present):
        ext_df = pd.DataFrame({
            f"EXT_SOURCE_{i}": uploaded_df[f"EXT_SOURCE_{i}"].values
            if f"EXT_SOURCE_{i}" in uploaded_df.columns
            else np.full(len(uploaded_df), medians.get(f"EXT_SOURCE_{i}", np.nan))
            for i in range(1, 4)
        })
        out["EXT_SOURCE_MEAN"] = ext_df.mean(axis=1).values
        out["EXT_SOURCE_MIN"]  = ext_df.min(axis=1).values
        out["EXT_SOURCE_MAX"]  = ext_df.max(axis=1).values
        out["EXT_SOURCE_STD"]  = ext_df.std(axis=1).values

    if "CODE_GENDER" in uploaded_df.columns:
        out["CODE_GENDER_M"] = (uploaded_df["CODE_GENDER"].str.upper() == "M").astype(float).values
    for src, dst in [("FLAG_OWN_CAR", "FLAG_OWN_CAR_Y"), ("FLAG_OWN_REALTY", "FLAG_OWN_REALTY_Y")]:
        if src in uploaded_df.columns and dst in feature_cols:
            out[dst] = (uploaded_df[src].astype(str).str.upper() == "Y").astype(float).values

    return out.astype(float)


# ─── Бизнес-правила ──────────────────────────────────────────────────────────

def check_business_rules(income: float, credit: float, annuity: float) -> list[str]:
    """Возвращает список нарушенных жёстких правил (пусто = нарушений нет).

    AMT_INCOME_TOTAL и AMT_ANNUITY — одного масштаба (месячные),
    поэтому DTI = annuity / income сравнивается напрямую.
    """
    violations: list[str] = []
    if income <= 0:
        return violations
    dti = annuity / income
    cti = credit  / income
    if dti >= 1.0:
        violations.append(
            f"Платёж {annuity:,.0f} ≥ дохода {income:,.0f} сом "
            f"(долговая нагрузка {dti:.0%} — невозможно обслуживать)"
        )
    elif dti >= 0.80:
        violations.append(
            f"Платёж {annuity:,.0f} составляет {dti:.0%} дохода {income:,.0f} сом "
            f"(критически высокая нагрузка, норма ≤ 30%)"
        )
    if cti > 240:
        violations.append(
            f"Кредит {credit:,.0f} = {cti:.0f}× месячного дохода "
            f"(норма: до 240×, т.е. 20 лет)"
        )
    return violations


# ─── Зоны риска ──────────────────────────────────────────────────────────────

def get_zone(prob: float, threshold: float = 0.5) -> tuple[str, str, str, str, str]:
    """Определяет зону риска относительно порога конкретной модели."""
    thr = max(threshold, 1e-6)
    if prob >= thr:
        excess = (prob - thr) / max(1.0 - thr, 1e-6)
        idx = 3 if excess < 0.20 else (4 if excess < 0.50 else 5)
    else:
        closeness = prob / thr
        idx = 0 if closeness < 0.25 else (1 if closeness < 0.65 else 2)
    return tuple(ZONE_DATA[idx])


def risk_level_label(prob: float, threshold: float) -> str:
    return get_zone(prob, threshold)[1]


# ─── Экспорт ─────────────────────────────────────────────────────────────────

def generate_excel_report(result_df: pd.DataFrame, threshold: float, model_name: str) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb  = Workbook()
    ws1 = wb.active
    ws1.title = "Результаты"

    green_fill = PatternFill(start_color="C6EFCE", fill_type="solid")
    red_fill   = PatternFill(start_color="FFC7CE", fill_type="solid")
    hdr_fill   = PatternFill(start_color="1F4E79", fill_type="solid")
    hdr_font   = Font(color="FFFFFF", bold=True)
    gray_font  = Font(color="888888", size=9)

    rus_row  = ["ID заявки", "Месячный доход", "Кредит", "Платёж/мес",
                "Риск (%)", "Уровень риска", "Решение"]
    tech_row = ["SK_ID_CURR", "AMT_INCOME_TOTAL", "AMT_GOODS_PRICE",
                "AMT_ANNUITY", "default_prob", "risk_level", "decision"]

    ws1.append(rus_row)
    for cell in ws1[1]:
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")
    ws1.append(tech_row)
    for cell in ws1[2]:
        cell.font = gray_font

    avail = [c for c in tech_row if c in result_df.columns]
    for row_vals in result_df[avail].itertuples(index=False):
        ws1.append(list(row_vals))
    for row in ws1.iter_rows(min_row=3):
        fill = red_fill if row[-1].value == "Отказать" else green_fill
        for cell in row:
            cell.fill = fill

    ws2 = wb.create_sheet("Статистика")
    n_ok = (result_df["decision"] == "Одобрить").sum()
    n_no = (result_df["decision"] == "Отказать").sum()
    for s in [
        ["Показатель", "Значение"],
        ["Всего заявок", len(result_df)],
        ["Одобрено", f"{n_ok} ({n_ok/len(result_df):.1%})"],
        ["Отказано",  f"{n_no} ({n_no/len(result_df):.1%})"],
        ["Средний риск", f"{result_df['default_prob'].mean():.1%}"],
        ["Дата оценки", datetime.now().strftime("%d.%m.%Y %H:%M")],
    ]:
        ws2.append(s)

    ws3 = wb.create_sheet("О модели")
    for s in [
        ["Параметр", "Значение"],
        ["Модель", model_name.upper()],
        ["Порог отсечения", f"{threshold:.4f}"],
        ["Метрика оптимизации", "F₂-score (β=2)"],
        ["Принцип", "Recall важнее Precision — снижение пропущенных дефолтов"],
        ["Дата оценки", datetime.now().strftime("%d.%m.%Y %H:%M")],
    ]:
        ws3.append(s)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def make_template_csv() -> str:
    return pd.DataFrame({
        "AMT_INCOME_TOTAL": [150000, 90000],
        "AMT_CREDIT":       [450000, 300000],
        "AMT_ANNUITY":      [22500,  18000],
        "AGE_YEARS":        [35, 28],
        "CODE_GENDER":      ["F", "M"],
        "FLAG_OWN_CAR":     ["N", "Y"],
        "FLAG_OWN_REALTY":  ["Y", "N"],
        "CNT_CHILDREN":     [0, 1],
        "CNT_FAM_MEMBERS":  [2, 3],
        "YEARS_EMPLOYED":   [5, 1],
        "IS_UNEMPLOYED":    [0, 0],
        "EXT_SOURCE_1":     [0.60, 0.30],
        "EXT_SOURCE_2":     [0.75, 0.35],
        "EXT_SOURCE_3":     [0.65, 0.40],
    }).to_csv(index=False, encoding="utf-8-sig")
